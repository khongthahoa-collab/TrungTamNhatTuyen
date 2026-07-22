from flask import render_template, redirect, url_for, flash, request, jsonify, abort
from flask_login import login_required, current_user
from datetime import date, datetime
from sqlalchemy import extract, func, case
from sqlalchemy.orm import joinedload
from extensions import db
from models import (TuitionPayment, Student, Class, Salary, Teacher, Course,
                    Expense, ExpenseCategory, TuitionMethod, Enrollment)
from blueprints.admin import admin_bp, require_admin, require_master
from services.salary_service import calculate_all_salaries, get_or_create_salary
from services.zalo_service import ZaloService
from services.tuition_service import (create_tuition_payment, record_payment, get_previous_month_debt,
                                      record_fee_adjustment, batch_previous_month_debts,
                                      void_tuition_payment, unvoid_tuition_payment, reverse_payment)
from services.academic_year_service import FrozenPeriodError, is_period_writable, list_academic_year_months
from blueprints.pagination_utils import paginate_list


# ────────────────────────────────────────────────────────────────
# Tuition
# ────────────────────────────────────────────────────────────────

def _tuition_overview_aggregate(month, year, class_id=None, course_id=None):
    """Shared by the /admin/tuition web view and the /api/v1/tuition/overview
    JSON endpoint, so the two can't drift. Tổng hợp bằng SQL, không tải toàn
    bộ bản ghi học phí của tháng vào Python (tránh chậm/502 khi dữ liệu lớn
    dần). Lists every active class (not just ones with tuition already
    generated) so a class with zero rows this month still shows up, with
    its live enrollment count, for the per-class "Tạo học phí" action.
    Returns (classes, class_summaries, total_collected, total_outstanding,
    total_expected)."""
    classes_q = Class.query.filter_by(is_active=True)
    if course_id:
        classes_q = classes_q.filter_by(course_id=course_id)
    if class_id:
        classes_q = classes_q.filter_by(id=class_id)
    classes = classes_q.order_by(Class.name).all()
    class_ids = [c.id for c in classes]

    enrolled_counts = {}
    if class_ids:
        enrolled_counts = dict(
            db.session.query(Enrollment.class_id, func.count(Enrollment.id))
            .filter(Enrollment.class_id.in_(class_ids), Enrollment.is_active == True)
            .group_by(Enrollment.class_id)
            .all()
        )

    # total counts every row, including voided ones — a voided bill still
    # occupies the (student, class, month, year) unique-index slot, so it
    # must keep counting as "generated"/"billed" or missing_count would
    # wrongly flag that student as needing a fresh bill (which
    # create_tuition_payment would then just refuse to create). Every
    # money/status figure below (paid/unpaid counts and amounts, carried
    # debt) excludes voided rows — that's the "excluded from revenue
    # reports" requirement.
    not_voided = TuitionPayment.is_voided == False
    total_due_expr = TuitionPayment.amount + TuitionPayment.debt_carried_over
    class_agg_query = (
        db.session.query(
            TuitionPayment.class_id,
            func.count(TuitionPayment.id).label('total'),
            func.sum(case((db.and_(TuitionPayment.is_paid == True, not_voided), 1), else_=0)).label('paid_count'),
            func.sum(case((db.and_(TuitionPayment.is_paid == False, not_voided), 1), else_=0)).label('unpaid_count'),
            func.coalesce(func.sum(case((not_voided, TuitionPayment.amount_collected), else_=0)), 0).label('collected_amount'),
            func.coalesce(func.sum(case(
                (db.and_(TuitionPayment.is_paid == False, not_voided), total_due_expr - TuitionPayment.amount_collected),
                else_=0,
            )), 0).label('outstanding_amount'),
            func.coalesce(func.sum(case((not_voided, TuitionPayment.debt_carried_over), else_=0)), 0).label('carried_debt_amount'),
        )
        .filter(TuitionPayment.month == month, TuitionPayment.year == year)
        .group_by(TuitionPayment.class_id)
    )
    if class_id:
        class_agg_query = class_agg_query.filter(TuitionPayment.class_id == class_id)
    by_class = {row.class_id: row for row in class_agg_query.all()}

    class_summaries = []
    for cls in classes:
        row = by_class.get(cls.id)
        enrolled_count = enrolled_counts.get(cls.id, 0)
        total = row.total if row else 0
        class_summaries.append({
            'class':          cls,
            'enrolled_count': enrolled_count,
            'generated':      row is not None,
            # Enrolled but not yet billed — e.g. a student added to the
            # class after tuition was already generated for this month.
            'missing_count':  max(0, enrolled_count - total),
            'total':          total,
            'paid_count':     row.paid_count if row else 0,
            'unpaid_count':   row.unpaid_count if row else 0,
            'paid_amount':    row.collected_amount if row else 0,
            'unpaid_amount':  row.outstanding_amount if row else 0,
            'carried_debt':   row.carried_debt_amount if row else 0,
        })

    total_collected = sum(r['paid_amount'] for r in class_summaries)
    total_outstanding = sum(r['unpaid_amount'] for r in class_summaries)
    total_expected = total_collected + total_outstanding
    return classes, class_summaries, total_collected, total_outstanding, total_expected


@admin_bp.route('/tuition')
@login_required
@require_admin
def tuition():
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year = request.args.get('year', today.year, type=int)
    class_id = request.args.get('class_id', type=int)
    course_id = request.args.get('course_id', type=int)
    not_generated = request.args.get('not_generated') == '1'

    classes, class_summaries, total_collected, total_outstanding, total_expected = \
        _tuition_overview_aggregate(month, year, class_id, course_id)

    if not_generated:
        # Lớp chưa tạo học phí tháng này — total=0 nghĩa là chưa có hoá đơn
        # nào (khác missing_count, vốn đếm học sinh thiếu hoá đơn ở 1 lớp
        # ĐÃ tạo học phí một phần).
        class_summaries = [s for s in class_summaries if not s['generated']]

    # Footer totals must reflect every class this month, not just the
    # current page — computed from the full list before slicing it.
    total_students = sum(r['total'] for r in class_summaries)
    total_paid_count = sum(r['paid_count'] for r in class_summaries)
    total_unpaid_count = sum(r['unpaid_count'] for r in class_summaries)
    total_carried_debt = sum(r['carried_debt'] for r in class_summaries)

    page = request.args.get('page', 1, type=int)
    summaries_pagination = paginate_list(class_summaries, page, per_page=10)

    courses = Course.query.filter_by(is_active=True).order_by(Course.name).all()
    # Newest-first; bounds both the dropdown and prev/next nav to periods
    # that actually belong to a registered academic year, instead of
    # unconditional +/-1 month arithmetic that could wander outside any
    # known year.
    year_months = list_academic_year_months()
    try:
        idx = year_months.index((year, month))
        prev_period = year_months[idx + 1] if idx + 1 < len(year_months) else None
        next_period = year_months[idx - 1] if idx > 0 else None
    except ValueError:
        prev_period = next_period = None

    return render_template('admin/finance/tuition.html',
                           classes=classes,
                           courses=courses,
                           class_summaries=summaries_pagination.items,
                           pagination=summaries_pagination,
                           total_students=total_students,
                           total_paid_count=total_paid_count,
                           total_unpaid_count=total_unpaid_count,
                           total_carried_debt=total_carried_debt,
                           month=month,
                           year=year,
                           selected_class_id=class_id,
                           selected_course_id=course_id,
                           not_generated=not_generated,
                           total_collected=total_collected,
                           total_outstanding=total_outstanding,
                           total_expected=total_expected,
                           year_months=year_months,
                           prev_period=prev_period,
                           next_period=next_period,
                           is_writable=is_period_writable(month, year),
                           today=today)


@admin_bp.route('/tuition/class/<int:class_id>')
@login_required
@require_admin
def tuition_class_detail(class_id):
    """Chi tiết học phí của một lớp trong một tháng"""
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year  = request.args.get('year',  today.year,  type=int)
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '').strip()

    cls = Class.query.get_or_404(class_id)

    base_query = (
        TuitionPayment.query
        .filter_by(class_id=class_id, month=month, year=year)
        .join(Student, TuitionPayment.student_id == Student.id)
        .options(joinedload(TuitionPayment.student))
    )
    if q:
        base_query = base_query.filter(Student.full_name.ilike(f'%{q}%'))
    if status_filter == 'paid':
        base_query = base_query.filter(TuitionPayment.is_paid == True, TuitionPayment.is_voided == False)
    elif status_filter == 'partial':
        base_query = base_query.filter(TuitionPayment.is_paid == False, TuitionPayment.is_voided == False,
                                       TuitionPayment.amount_collected > 0)
    elif status_filter == 'unpaid':
        base_query = base_query.filter(TuitionPayment.is_paid == False, TuitionPayment.is_voided == False,
                                       TuitionPayment.amount_collected == 0)
    elif status_filter == 'voided':
        base_query = base_query.filter(TuitionPayment.is_voided == True)
    base_query = base_query.order_by(TuitionPayment.is_paid, Student.full_name)
    pagination = base_query.paginate(page=page, per_page=10, error_out=False)
    records = pagination.items

    # Danh sách đầy đủ (không phân trang) chỉ dùng để dựng ảnh "Xuất ảnh học
    # phí nhóm" — thẻ ảnh phải liệt kê đủ cả lớp, không chỉ 10 dòng của
    # trang hiện tại. Loại hoá đơn đã hủy, không xuất hiện trên thông báo.
    all_records = (
        TuitionPayment.query
        .filter_by(class_id=class_id, month=month, year=year, is_voided=False)
        .join(Student, TuitionPayment.student_id == Student.id)
        .options(joinedload(TuitionPayment.student))
        .order_by(Student.full_name)
        .all()
    )

    from models import BankAccount
    from services.tuition_service import build_vietqr_url
    active_bank_accounts = BankAccount.query.filter_by(is_active=True).order_by(BankAccount.id).all()
    # 1 QR sẵn cho từng tài khoản đang bật — JS đổi hiển thị khi admin chọn
    # tài khoản khác trên thẻ ảnh (nhóm/cá nhân), không cần gọi lại server.
    group_qr_by_account = {
        acc.id: build_vietqr_url(acc.bank_id, acc.account_number, acc.account_name,
                                 add_info=f'HP lop {cls.name} thang {month:02d} nam {year}')
        for acc in active_bank_accounts
    }

    # KPI totals exclude voided bills entirely (excluded from revenue
    # reports) — but the row list above (base_query) still shows them, and
    # missing_count below still counts them as "billed" (a voided bill
    # keeps its unique-index slot, so it shouldn't look like it needs a
    # fresh one).
    total_due_expr = TuitionPayment.amount + TuitionPayment.debt_carried_over
    total, paid_count, unpaid_count, collected_amount, outstanding_amount = (
        db.session.query(
            func.count(TuitionPayment.id),
            func.sum(case((TuitionPayment.is_paid == True, 1), else_=0)),
            func.sum(case((TuitionPayment.is_paid == False, 1), else_=0)),
            func.coalesce(func.sum(TuitionPayment.amount_collected), 0),
            func.coalesce(func.sum(case(
                (TuitionPayment.is_paid == False, total_due_expr - TuitionPayment.amount_collected),
                else_=0,
            )), 0),
        )
        .filter_by(class_id=class_id, month=month, year=year, is_voided=False)
        .first()
    )

    # Students enrolled in the class but with no TuitionPayment row this
    # month — happens whenever a student is added to the class *after*
    # tuition was already generated for that month (monthly_fee_generate
    # itself already skips students who already have a row, so it's safe
    # to re-run; the gap was that there was previously no way to trigger
    # it again once a class had any tuition at all — see the "Tạo học
    # phí cho N học sinh còn thiếu" button below).
    enrolled_ids = {e.student_id for e in Enrollment.query.filter_by(class_id=class_id, is_active=True).all()}
    billed_ids = {t.student_id for t in TuitionPayment.query.filter_by(
        class_id=class_id, month=month, year=year).all()}
    missing_count = len(enrolled_ids - billed_ids)

    return render_template('admin/finance/tuition_class_detail.html',
                           cls=cls, records=records, all_records=all_records, pagination=pagination,
                           total=total or 0, paid_count=paid_count or 0,
                           unpaid_count=unpaid_count or 0,
                           collected_amount=collected_amount or 0,
                           outstanding_amount=outstanding_amount or 0,
                           missing_count=missing_count,
                           is_writable=is_period_writable(month, year),
                           active_bank_accounts=active_bank_accounts,
                           group_qr_by_account=group_qr_by_account,
                           build_vietqr_url=build_vietqr_url,
                           q=q, status_filter=status_filter,
                           month=month, year=year, today=today)


@admin_bp.route('/tuition/class/<int:class_id>/export.xlsx')
@login_required
@require_admin
def tuition_class_export(class_id):
    """Xuất Excel toàn bộ danh sách học phí của lớp trong tháng — không giới
    hạn theo trang (khác với bảng hiển thị trên màn hình, vốn phân trang
    50/trang)."""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font

    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year = request.args.get('year', today.year, type=int)

    cls = Class.query.get_or_404(class_id)
    records = (
        TuitionPayment.query
        .filter_by(class_id=class_id, month=month, year=year)
        .join(Student, TuitionPayment.student_id == Student.id)
        .options(joinedload(TuitionPayment.student))
        .order_by(TuitionPayment.is_paid, Student.full_name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f'HP T{month}-{year}'
    ws.append([f'Học phí {cls.name} — Tháng {month}/{year}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    headers = ['STT', 'Học sinh', 'SĐT phụ huynh', 'Học phí tháng này', 'Nợ cũ',
              'Tổng cộng cần thu', 'Đã đóng', 'Trạng thái', 'Ghi chú']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    for i, r in enumerate(records, start=1):
        ws.append([
            i, r.student.full_name, r.student.parent_phone or '',
            r.amount, r.debt_carried_over or 0, r.total_due, r.amount_collected or 0,
            r.status_label, r.note or '',
        ])

    for col, width in zip('ABCDEFGHI', (5, 24, 14, 16, 14, 18, 14, 14, 26)):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'hoc_phi_{cls.name}_{month}_{year}.xlsx'.replace(' ', '_')
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/tuition/add', methods=['POST'])
@login_required
@require_admin
def tuition_add():
    student_id = request.form.get('student_id', type=int)
    class_id = request.form.get('class_id', type=int)
    amount = request.form.get('amount', type=float)
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    note = request.form.get('note', '').strip()

    if not all([student_id, class_id, amount, month, year]):
        flash('Vui lòng điền đầy đủ thông tin.', 'danger')
        return redirect(url_for('admin.tuition'))

    try:
        tp, created = create_tuition_payment(student_id, class_id, month, year, amount, note=note)
    except (FrozenPeriodError, ValueError) as e:
        flash(str(e), 'danger')
        return redirect(url_for('admin.tuition', month=month, year=year))

    if not created:
        flash('Đã có bản ghi học phí tháng này rồi.', 'warning')
    else:
        flash('Đã thêm bản ghi học phí.', 'success')
        db.session.commit()
    return redirect(url_for('admin.tuition', month=month, year=year))


@admin_bp.route('/tuition/<int:payment_id>/mark-paid', methods=['POST'])
@login_required
@require_admin
def tuition_mark_paid(payment_id):
    method = request.form.get('method', 'cash')
    amount = request.form.get('amount', type=float)
    note = request.form.get('note', '').strip() or None
    try:
        tp = record_payment(payment_id, amount, method, current_user.id, note=note)
    except FrozenPeriodError as e:
        flash(str(e), 'danger')
        return redirect(request.referrer or url_for('admin.tuition'))
    if not tp:
        abort(404)
    flash(f'Đã ghi nhận thanh toán học phí cho {tp.student.full_name}.', 'success')
    return redirect(request.referrer or url_for('admin.tuition', month=tp.month, year=tp.year))


@admin_bp.route('/tuition/zalo-remind', methods=['POST'])
@login_required
@require_admin
def tuition_remind_zalo():
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    class_id = request.form.get('class_id', type=int)

    from models import BankAccount
    active_acc = BankAccount.query.filter_by(is_active=True).order_by(BankAccount.id).first()
    bank_info = f'{active_acc.bank_name} – {active_acc.account_number} – {active_acc.account_name}' if active_acc else ''

    query = TuitionPayment.query.filter_by(month=month, year=year, is_paid=False)
    if class_id:
        query = query.filter_by(class_id=class_id)

    records = query.all()
    sent = 0
    for tp in records:
        ZaloService.send_tuition_reminder(tp.student, tp.class_, month, year, tp.amount, bank_info)
        sent += 1

    flash(f'Đã gửi {sent} thông báo nhắc học phí.', 'success')
    return redirect(url_for('admin.tuition', month=month, year=year))


# ────────────────────────────────────────────────────────────────
# Expenses
# ────────────────────────────────────────────────────────────────

@admin_bp.route('/expenses')
@login_required
@require_admin
def expenses():
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year = request.args.get('year', today.year, type=int)
    category = request.args.get('category', '')
    page = request.args.get('page', 1, type=int)

    query = Expense.query.filter(
        db.extract('month', Expense.expense_date) == month,
        db.extract('year', Expense.expense_date) == year,
    )
    if category:
        query = query.filter_by(category=category)

    total, tax_deductible = query.with_entities(
        func.coalesce(func.sum(Expense.amount), 0),
        func.coalesce(func.sum(case((Expense.is_tax_deductible == True, Expense.amount), else_=0)), 0),
    ).first()

    pagination = query.order_by(Expense.expense_date.desc()).paginate(page=page, per_page=50, error_out=False)
    records = pagination.items

    is_filtered = bool(category or month != today.month or year != today.year)

    return render_template('admin/finance/expenses.html',
                           records=records,
                           pagination=pagination,
                           month=month,
                           year=year,
                           selected_category=category,
                           total=total,
                           tax_deductible=tax_deductible,
                           categories=ExpenseCategory.LABELS,
                           is_filtered=is_filtered,
                           today=today)


@admin_bp.route('/expenses/add', methods=['POST'])
@login_required
@require_admin
def expense_add():
    category = request.form.get('category', '')
    amount = request.form.get('amount', type=float)
    description = request.form.get('description', '').strip()
    date_str = request.form.get('expense_date', '')
    is_tax = request.form.get('is_tax_deductible') == '1'
    receipt_no = request.form.get('receipt_no', '').strip()

    if not category or not amount:
        flash('Vui lòng nhập danh mục và số tiền.', 'danger')
        return redirect(url_for('admin.expenses'))

    today = date.today()
    try:
        expense_date = date.fromisoformat(date_str) if date_str else today
    except ValueError:
        expense_date = today

    exp = Expense(
        category=category,
        amount=amount,
        description=description,
        expense_date=expense_date,
        fiscal_year=expense_date.year,
        is_tax_deductible=is_tax,
        receipt_no=receipt_no,
        created_by=current_user.id,
    )
    db.session.add(exp)
    db.session.commit()
    flash('Đã ghi chi phí.', 'success')
    return redirect(url_for('admin.expenses',
                            month=expense_date.month, year=expense_date.year))


@admin_bp.route('/expenses/<int:exp_id>/delete', methods=['POST'])
@login_required
@require_admin
def expense_delete(exp_id):
    exp = Expense.query.get_or_404(exp_id)
    db.session.delete(exp)
    db.session.commit()
    flash('Đã xóa chi phí.', 'success')
    return redirect(request.referrer or url_for('admin.expenses'))


# ────────────────────────────────────────────────────────────────
# Salary
# ────────────────────────────────────────────────────────────────

@admin_bp.route('/salary')
@login_required
@require_master
def salary():
    from models import User, Schedule
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year = request.args.get('year', today.year, type=int)
    page = request.args.get('page', 1, type=int)

    pagination = (Teacher.query
                 .join(Teacher.user).filter(User.is_deleted == False)
                 .order_by(User.full_name)
                 .paginate(page=page, per_page=50, error_out=False))
    teachers = pagination.items
    teacher_ids = [t.id for t in teachers]
    salaries_by_teacher = {
        s.teacher_id: s for s in Salary.query.filter_by(month=month, year=year).all()
    }
    session_counts = {}
    sub_counts = {}
    if teacher_ids:
        session_counts = dict(
            db.session.query(Schedule.teacher_id, func.count(Schedule.id))
            .filter(
                Schedule.teacher_id.in_(teacher_ids),
                Schedule.substitute_teacher_id.is_(None),
                Schedule.is_cancelled == False,
                extract('month', Schedule.date) == month,
                extract('year', Schedule.date) == year,
            )
            .group_by(Schedule.teacher_id).all()
        )
        sub_counts = dict(
            db.session.query(Schedule.substitute_teacher_id, func.count(Schedule.id))
            .filter(
                Schedule.substitute_teacher_id.in_(teacher_ids),
                Schedule.is_cancelled == False,
                extract('month', Schedule.date) == month,
                extract('year', Schedule.date) == year,
            )
            .group_by(Schedule.substitute_teacher_id).all()
        )

    return render_template('admin/finance/salary.html',
                           teachers=teachers,
                           pagination=pagination,
                           salaries_by_teacher=salaries_by_teacher,
                           session_counts=session_counts,
                           sub_counts=sub_counts,
                           month=month,
                           year=year,
                           today=today)


@admin_bp.route('/salary/calculate', methods=['POST'])
@login_required
@require_master
def salary_calculate():
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    created = calculate_all_salaries(month, year)
    db.session.flush()

    total_amount = db.session.query(func.coalesce(func.sum(Salary.total), 0)).filter_by(
        month=month, year=year
    ).scalar()
    expense = Expense.query.filter(
        Expense.category == ExpenseCategory.SALARY,
        extract('month', Expense.expense_date) == month,
        extract('year', Expense.expense_date) == year,
    ).first()
    if expense:
        expense.amount = total_amount
    else:
        expense = Expense(
            category=ExpenseCategory.SALARY,
            amount=total_amount,
            description=f'Lương giáo viên tháng {month}/{year}',
            expense_date=date(year, month, 1),
            created_by=current_user.id,
        )
        db.session.add(expense)

    db.session.commit()
    flash(f'Đã tính lương cho {len(created)} giáo viên mới. '
          f'Tổng chi lương tháng {month}/{year}: {total_amount:,.0f} đ đã cập nhật vào Chi phí.', 'success')
    return redirect(url_for('admin.salary', month=month, year=year))


@admin_bp.route('/salary/start/<int:teacher_id>', methods=['POST'])
@login_required
@require_master
def salary_start(teacher_id):
    """Create-if-missing entry point for clicking a teacher's row on the
    salary list (both the name and the "Chưa tính lương" status lead here)."""
    teacher = Teacher.query.get_or_404(teacher_id)
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    salary, _ = get_or_create_salary(teacher, month, year)
    db.session.commit()
    return redirect(url_for('admin.salary_detail', salary_id=salary.id))


@admin_bp.route('/salary/detail/<int:salary_id>', methods=['GET', 'POST'])
@login_required
@require_master
def salary_detail(salary_id):
    sal = Salary.query.get_or_404(salary_id)

    if request.method == 'POST':
        sal.base_amount = request.form.get('base_amount', 0, type=float)
        sal.bonus = request.form.get('bonus', 0, type=float)
        sal.deduction = request.form.get('deduction', 0, type=float)
        sal.advance = request.form.get('advance', 0, type=float)
        total = request.form.get('total', type=float)
        sal.total = total if total is not None else (sal.base_amount + sal.bonus - sal.deduction - sal.advance)
        sal.note = request.form.get('note', '').strip()
        db.session.commit()
        flash(f'Đang chỉnh sửa thông tin cho Phiếu lương {sal.month}/{sal.year} - {sal.teacher.full_name}', 'warning')
        return redirect(url_for('admin.salary_detail', salary_id=sal.id))

    return render_template('admin/finance/salary_detail.html', salary=sal)


@admin_bp.route('/salary/print')
@login_required
@require_master
def salary_print():
    from models import User
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    salaries = (Salary.query.filter_by(month=month, year=year)
               .join(Salary.teacher).join(Teacher.user)
               .order_by(User.full_name).all())
    return render_template('admin/finance/salary_print.html',
                           salaries=salaries, month=month, year=year)




# ────────────────────────────────────────────────────────────────
# Monthly tuition generation — default fee is simply Class.monthly_fee ×
# each actively-enrolled student; no per-month proration config. (The old
# "Cấu hình học phí" week-proration feature and its MonthlyClassFee table
# were removed — this app no longer needs to bill a fraction of a month.
# A per-student exception for a given month is handled by editing that
# one row's fee directly on the class-detail page instead.)
# ────────────────────────────────────────────────────────────────

@admin_bp.route('/tuition/monthly/generate', methods=['POST'])
@login_required
@require_admin
def monthly_fee_generate():
    """Tạo hàng loạt TuitionPayment cho tất cả học sinh đang học trong một
    lớp cụ thể, dùng học phí chuẩn của lớp (Class.monthly_fee) làm mặc định.

    Luôn giới hạn theo một lớp — generate cho toàn bộ trường trong một
    request từng gây 502 khi số học sinh lớn (mỗi học sinh cần vài query
    độc lập: tính nợ cũ, kiểm tra tồn tại, insert). Bó buộc theo lớp giữ
    kích thước mỗi request nhỏ và cố định, bất kể trường có bao nhiêu lớp."""
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    class_id = request.form.get('class_id', type=int)
    # Re-running this for a class that already has tuition generated is
    # exactly how a class picks up students who were enrolled *after* the
    # first generate — the loop below already skips anyone who already has
    # a row, so this is always safe to call again. When triggered from the
    # class-detail page ("Tạo học phí cho N học sinh còn thiếu"), send the
    # admin back there instead of the overview.
    redirect_target = (
        url_for('admin.tuition_class_detail', class_id=class_id, month=month, year=year)
        if request.form.get('redirect_to') == 'class_detail' and class_id
        else url_for('admin.tuition', month=month, year=year)
    )
    if not class_id:
        flash('Vui lòng chọn một lớp để tạo học phí.', 'danger')
        return redirect(redirect_target)

    if not is_period_writable(month, year):
        flash('Không thể sửa đổi dữ liệu tài chính của năm học đã đóng băng', 'danger')
        return redirect(redirect_target)

    cls = Class.query.get_or_404(class_id)

    enrollments = Enrollment.query.filter_by(class_id=class_id, is_active=True).all()
    student_ids = [e.student_id for e in enrollments]
    existing_student_ids = set()
    if student_ids:
        existing_student_ids = {
            t.student_id for t in TuitionPayment.query.filter(
                TuitionPayment.class_id == class_id, TuitionPayment.student_id.in_(student_ids),
                TuitionPayment.month == month, TuitionPayment.year == year,
            ).all()
        }

    # Batch every student's previous-month debt in 1-2 queries instead of
    # one SELECT per student inside the loop — the N+1 pattern that
    # already caused a 502 once this session, just in this route instead.
    # skip_period_check=True below: already checked once above for the
    # whole batch, so create_tuition_payment() doesn't re-run the same
    # AcademicYear query per student (which would otherwise silently
    # cancel out this exact batching).
    to_create_ids = [sid for sid in student_ids if sid not in existing_student_ids]
    debt_map = batch_previous_month_debts(to_create_ids, class_id, month, year)

    fee = cls.monthly_fee or 0
    created = 0
    try:
        for student_id in to_create_ids:
            _, was_created = create_tuition_payment(
                student_id, class_id, month, year, fee,
                debt_override=debt_map.get(student_id, 0), skip_period_check=True)
            if was_created:
                created += 1
    except FrozenPeriodError as e:
        db.session.rollback()
        flash(str(e), 'danger')
        return redirect(redirect_target)

    db.session.commit()
    flash(f'Đã tạo {created} học phí mới cho lớp {cls.name} tháng {month}/{year}.', 'success')
    return redirect(redirect_target)


@admin_bp.route('/tuition/<int:payment_id>/adjust-amount', methods=['POST'])
@login_required
@require_admin
def tuition_adjust_amount(payment_id):
    """Admin điều chỉnh học phí tháng này cho một học sinh cụ thể — chỉ áp
    dụng cho tháng đang xem, ghi lại vào TuitionFeeAuditLog (record_fee_adjustment)."""
    tp = TuitionPayment.query.get_or_404(payment_id)
    if tp.is_paid:
        flash('Học phí đã được thanh toán, không thể chỉnh sửa số tiền.', 'danger')
        return redirect(url_for('admin.tuition', month=tp.month, year=tp.year))

    new_amount = request.form.get('amount', type=float)
    note = request.form.get('note', '').strip() or None
    if new_amount is not None:
        try:
            record_fee_adjustment(payment_id, new_amount, current_user.id, note=note)
            flash(f'Đã cập nhật học phí {tp.student.full_name}: {int(new_amount):,}₫.', 'success')
        except (FrozenPeriodError, ValueError) as e:
            flash(str(e), 'danger')
    return redirect(request.referrer or url_for('admin.tuition', month=tp.month, year=tp.year))


@admin_bp.route('/tuition/<int:payment_id>/void', methods=['POST'])
@login_required
@require_admin
def tuition_void(payment_id):
    """Hủy (soft-delete) một hoá đơn học phí lập nhầm — không xoá bản ghi,
    chỉ loại khỏi các báo cáo/tổng doanh thu. Bắt buộc nhập lý do."""
    tp = TuitionPayment.query.get_or_404(payment_id)
    reason = request.form.get('reason', '')
    try:
        void_tuition_payment(payment_id, reason, current_user.id)
        flash(f'Đã hủy học phí {tp.student.full_name} tháng {tp.month}/{tp.year}.', 'success')
    except (FrozenPeriodError, ValueError) as e:
        flash(str(e), 'danger')
    return redirect(request.referrer or url_for('admin.tuition', month=tp.month, year=tp.year))


@admin_bp.route('/tuition/<int:payment_id>/unvoid', methods=['POST'])
@login_required
@require_admin
def tuition_unvoid(payment_id):
    """Khôi phục một hoá đơn đã bị hủy nhầm."""
    tp = TuitionPayment.query.get_or_404(payment_id)
    try:
        unvoid_tuition_payment(payment_id)
        flash(f'Đã khôi phục học phí {tp.student.full_name} tháng {tp.month}/{tp.year}.', 'success')
    except (FrozenPeriodError, ValueError) as e:
        flash(str(e), 'danger')
    return redirect(request.referrer or url_for('admin.tuition', month=tp.month, year=tp.year))


@admin_bp.route('/tuition/<int:payment_id>/reverse-payment', methods=['POST'])
@login_required
@require_admin
def tuition_reverse_payment(payment_id):
    """Hoàn tác một khoản thu đã ghi nhận nhầm (sai số tiền, nhầm học
    sinh…) — khác với "Hủy" (hủy hoá đơn chưa thu tiền). Không sửa trực
    tiếp amount_collected, mà ghi một giao dịch bù trừ âm vào sổ quỹ
    (reverse_payment)."""
    tp = TuitionPayment.query.get_or_404(payment_id)
    reason = request.form.get('reason', '')
    try:
        reverse_payment(payment_id, reason, current_user.id)
        flash(f'Đã hoàn tác thanh toán của {tp.student.full_name} tháng {tp.month}/{tp.year}.', 'success')
    except (FrozenPeriodError, ValueError) as e:
        flash(str(e), 'danger')
    return redirect(request.referrer or url_for('admin.tuition', month=tp.month, year=tp.year))
